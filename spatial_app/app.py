from fastapi import FastAPI, UploadFile, File, Request, Form
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openai import OpenAI
from openpyxl import Workbook, load_workbook
from typing import List
import os
import io
import json
import re
import csv

from rubrics import rubric3, spatial_words

app = FastAPI()
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

client = OpenAI(api_key=os.environ.get("OPENAI_API_KEY"))

rubric = json.dumps(rubric3, indent=2)
spatial_dict = json.dumps(spatial_words)

MODEL = "gpt-4o-mini"
BATCH_SIZE = 8
TEMPERATURE = 0

admin_system_prompt = f"""You are the Admin Agent.
Your role is to receive an utterance from a parent-child conversation,
send this information to the Judger agents for analysis,
aggregate the Judger agents' decisions,
and send the aggregated decisions to the Critic agent.
"""

judger_system_prompt = f"""
You are a Judger Agent specializing in analyzing utterances for spatial language.

You will receive a **full transcript** as a JSON array of utterance objects.

Each object has:
- "lineNo"
- "speaker"
- "utterance"

Follow the below procedure for EACH utterance within the transcript.

**Your Task:**

1. **Identify Spatial Words from the Dictionary:**
   - Carefully examine the given utterance.
   - **Check every word** in the utterance against the **spatial dictionary** provided.
   - **List all words** from the utterance that appear in the dictionary.

2. **Assess Usage According to the Rubric:**
   - For each word identified from the dictionary, **determine if it is used in a spatial context** as defined in the rubric and its appendix.
   - Refer to the rubric to understand how spatial words should be used to be considered spatial language.

3. **Identify Additional Spatial Words According to the Rubric:**
   - If no words from the dictionary are found, or if you suspect other words might be spatial, **carefully analyze the utterance for any other spatial words or phrases** according to the rubric.
   - **Consider synonyms or contextually spatial terms** that may not be in the dictionary but fit the rubric's criteria.

4. **Critically Review Your Analysis:**
   - **Double-check** your findings to ensure you have not **missed any spatial words**.
   - **Ensure accuracy** by confirming that all identified words truly meet the spatial criteria without misidentifying non-spatial words.

**Instructions for Your Response:**

- **Provide a detailed reasoning** for your decisions, referencing both the dictionary and the rubric.
- **Conclude with a clear 'yes' or 'no' decision** on whether the utterance contains spatial language.
- **List all spatial words identified**, if any.
- **Use the 'provide_judgment' function** to deliver your response.

**Resources Provided:**

- **Rubric:** {rubric}
- **Spatial Dictionary:** {spatial_dict}

**Important Notes:**

- **Be meticulous** in your analysis to avoid missing any spatial words.
- **Avoid misidentifying** words that are not spatial according to the rubric.
- **Your critical thinking and attention to detail are crucial** for accurate results.

Output format:
Return JSONL ONLY.
One JSON object per input line.
No markdown fences.
No commentary outside the JSONL.

Each JSON object must have exactly:
{{
  "lineNo": number,
  "decision": "yes" or "no",
  "spatialWordsDetected": string[],
  "justification": string
}}
"""

judger_personalities = [
    f"""You are a trained coder in spatial cognition with exceptional attention to detail and precision in identification and accuracy.
{judger_system_prompt}""",

    f"""You are a cognitive scientist specializing in spatial cognition and spatial thinking.
{judger_system_prompt}""",

    f"""You are a linguist specialized in discerning the meaning of words in context, with a focus on spatial language.
{judger_system_prompt}"""
]

critic_system_prompt = f"""
You are the Critic Agent tasked with evaluating the decisions made by the Judger agents.

**Your Task:**

1. **Review Judger Decisions Critically:**
   - **Examine each Judger's reasoning**, decisions, and the spatial words they identified.
   - **Verify the accuracy** of their analyses by referencing the **spatial dictionary** and **rubric** provided.

2. **Ensure No Spatial Words Are Missed or Misidentified:**
   - **Identify any spatial words** that may have been **missed** by the Judger agents.
   - **Correct any misidentified words**, ensuring that only words fitting the rubric's criteria are considered spatial.

3. **Provide a Final Decision:**
   - **Aggregate the Judger agents' decisions** to form a cohesive final decision.
   - **Provide a clear 'yes' or 'no' decision** on whether the utterance contains spatial language.
   - **List all spatial words identified**, ensuring accuracy and completeness.
   - **Justify your final decision** with detailed reasoning, citing specific parts of the rubric and dictionary as necessary.

**Instructions for Your Response:**

- **Deliver your response using the 'aggregate_decisions' function.**
- **Include detailed justification** for your decision to demonstrate thorough critical analysis.

**Resources Provided:**

- **Rubric:** {rubric}
- **Spatial Dictionary:** {spatial_dict}

**Important Notes:**

- **Be thorough and unbiased** in your review.
- **Ensure that no spatial words are overlooked**.
- **Avoid including words that do not meet the spatial criteria** as defined by the rubric.

You will receive:
1. The complete transcript.
2. The outputs from multiple Judger agents.

For EACH lineNo in the transcript, compare the Judgers' decisions and reasoning.
Produce ONE final judgment for each lineNo.

You must call the aggregate_decisions function with one judgment per transcript line.
Return exactly one final judgment per input lineNo.
Do not skip lines.
Do not summarize the whole transcript.
"""

judgment_tools = [
    {
        "type": "function",
        "function": {
            "name": "provide_judgment",
            "description": "Provide one spatial-language judgment for each utterance in a transcript.",
            "parameters": {
                "type": "object",
                "properties": {
                    "judgments": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "lineNo": {"type": "number"},
                                "decision": {"type": "string", "enum": ["yes", "no"]},
                                "spatialWordsDetected": {
                                    "type": "array",
                                    "items": {"type": "string"}
                                },
                                "justification": {"type": "string"}
                            },
                            "required": [
                                "lineNo",
                                "decision",
                                "spatialWordsDetected",
                                "justification"
                            ],
                            "additionalProperties": False
                        }
                    }
                },
                "required": ["judgments"],
                "additionalProperties": False
            }
        }
    }
]

admin_tools = [
    {
        "type": "function",
        "function": {
            "name": "send_to_judgers",
            "description": "Sends the utterance to the Judger agents for analysis.",
            "parameters": {
                "type": "object",
                "properties": {
                    "utterance": {"type": "string", "description": "The sentence to analyze."}
                  },
                "required": ["utterance"],
                "additionalProperties": False
            }
        },
        "strict": True
    }
]

critic_tools = [
    {
        "type": "function",
        "function": {
            "name": "aggregate_decisions",
            "description": "Aggregates Judger outputs and provides one final spatial-language judgment for each utterance in a transcript.",
            "parameters": {
                "type": "object",
                "properties": {
                    "judgments": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "lineNo": {"type": "number"},
                                "decision": {
                                    "type": "string",
                                    "enum": ["yes", "no"]
                                },
                                "spatialWordsDetected": {
                                    "type": "array",
                                    "items": {"type": "string"}
                                },
                                "justification": {
                                    "type": "string"
                                }
                            },
                            "required": [
                                "lineNo",
                                "decision",
                                "spatialWordsDetected",
                                "justification"
                            ],
                            "additionalProperties": False
                        }
                    }
                },
                "required": ["judgments"],
                "additionalProperties": False
            }
        }
    }
]


def extract_tool_judgments(resp):
    msg = resp.choices[0].message
    tool_calls = msg.tool_calls or []

    for tc in tool_calls:
        fn_name = tc.function.name
        args = json.loads(tc.function.arguments)

        if fn_name == "provide_judgment":
            return args.get("judgments", [])

        if fn_name == "aggregate_decisions":
            return args.get("judgments", [])

    text = msg.content or ""
    return parse_jsonl(text)


def parse_txt(text: str):
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    return [
        {"lineNo": i + 1, "speaker": "", "utterance": ln}
        for i, ln in enumerate(lines)
    ]


def cleanup_chat_utterance(s: str) -> str:
    out = str(s)
    out = re.sub(r"\[[^\]]*\]", "", out)
    out = re.sub(r"<([^>]+)>", r"\1", out)
    out = re.sub(r"\s+", " ", out).strip()
    return out


def parse_cha(text: str):
    raw_lines = text.splitlines()
    out = []
    current = None
    line_no = 0

    for raw in raw_lines:
        line = raw.replace("\ufeff", "")
        if not line.strip():
            continue
        if line.startswith("@"):
            continue

        m = re.match(r"^\*([A-Za-z0-9_]+):\s*(.*)$", line)
        if m:
            if current:
                out.append(current)
            line_no += 1
            current = {
                "lineNo": line_no,
                "speaker": m.group(1),
                "utterance": cleanup_chat_utterance(m.group(2) or "")
            }
            continue

        if raw.startswith(" ") and current:
            current["utterance"] += " " + cleanup_chat_utterance(line.strip())

    if current:
        out.append(current)

    return out


def parse_csv(text: str):
    reader = csv.DictReader(io.StringIO(text))
    rows = []

    headers = reader.fieldnames or []
    lower_headers = {h.lower().strip(): h for h in headers}

    speaker_col = (
        lower_headers.get("speaker")
        or lower_headers.get("participant")
        or lower_headers.get("child")
        or lower_headers.get("role")
    )

    utterance_col = (
        lower_headers.get("utterance")
        or lower_headers.get("transcript")
        or lower_headers.get("text")
        or lower_headers.get("sentence")
        or lower_headers.get("line")
    )

    if not utterance_col:
        raise ValueError("CSV must have an utterance/text/transcript/sentence/line column.")

    for i, row in enumerate(reader, start=1):
        utterance = str(row.get(utterance_col, "")).strip()
        if not utterance:
            continue

        rows.append({
            "lineNo": i,
            "speaker": str(row.get(speaker_col, "")).strip() if speaker_col else "",
            "utterance": utterance
        })

    return rows


def parse_xlsx(file_bytes: bytes):
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    data = list(ws.iter_rows(values_only=True))
    if not data:
        return []

    headers = [str(x or "").strip() for x in data[0]]
    lower_headers = {h.lower(): idx for idx, h in enumerate(headers)}

    speaker_idx = (
        lower_headers.get("speaker")
        if "speaker" in lower_headers else
        lower_headers.get("participant")
        if "participant" in lower_headers else
        lower_headers.get("child")
        if "child" in lower_headers else
        lower_headers.get("role")
    )

    utterance_idx = (
        lower_headers.get("utterance")
        if "utterance" in lower_headers else
        lower_headers.get("transcript")
        if "transcript" in lower_headers else
        lower_headers.get("text")
        if "text" in lower_headers else
        lower_headers.get("sentence")
        if "sentence" in lower_headers else
        lower_headers.get("line")
    )

    if utterance_idx is None:
        raise ValueError("XLSX must have an utterance/text/transcript/sentence/line column.")

    rows = []
    for i, row in enumerate(data[1:], start=2):
        utterance = str(row[utterance_idx] or "").strip()
        if not utterance:
            continue

        speaker = ""
        if speaker_idx is not None and speaker_idx < len(row):
            speaker = str(row[speaker_idx] or "").strip()

        rows.append({
            "lineNo": i,
            "speaker": speaker,
            "utterance": utterance
        })

    return rows


def chunk_rows(rows, size):
    return [rows[i:i + size] for i in range(0, len(rows), size)]


def strip_code_fences(text: str) -> str:
    s = str(text or "").strip()
    s = re.sub(r"^```jsonl\s*", "", s, flags=re.IGNORECASE)
    s = re.sub(r"^```json\s*", "", s, flags=re.IGNORECASE)
    s = re.sub(r"^```\s*", "", s, flags=re.IGNORECASE)
    s = re.sub(r"```$", "", s, flags=re.IGNORECASE)
    return s.strip()


def parse_jsonl(text: str):
    s = strip_code_fences(text)
    lines = [ln.strip() for ln in s.splitlines() if ln.strip()]
    out = []

    for line in lines:
        if not line.startswith("{"):
            continue
        try:
            out.append(json.loads(line))
        except json.JSONDecodeError:
            continue

    return out


def normalize_items(items, batch):
    by_line = {item.get("lineNo"): item for item in items if isinstance(item, dict)}
    normalized = []

    for row in batch:
        item = by_line.get(row["lineNo"])
        if not item:
            normalized.append({
                "lineNo": row["lineNo"],
                "decision": "no",
                "spatialWordsDetected": [],
                "justification": "No valid output returned for this line."
            })
            continue

        normalized.append({
            "lineNo": row["lineNo"],
            "decision": "yes" if str(item.get("decision", "")).lower() == "yes" else "no",
            "spatialWordsDetected": item.get("spatialWordsDetected", []) if isinstance(item.get("spatialWordsDetected"), list) else [],
            "justification": item.get("justification", "") if isinstance(item.get("justification"), str) else ""
        })

    return normalized

def is_reasoning_model(model: str) -> bool:
    return model.startswith("gpt-5.5")


def convert_chat_tool_to_responses_tool(chat_tool):
    fn = chat_tool["function"]
    return {
        "type": "function",
        "name": fn["name"],
        "description": fn.get("description", ""),
        "parameters": fn["parameters"]
    }


def extract_response_tool_judgments(resp, tool_name: str):
    for item in resp.output:
        if getattr(item, "type", None) == "function_call" and getattr(item, "name", None) == tool_name:
            args = json.loads(item.arguments)
            return args.get("judgments", [])

    return []


def call_model_with_tool(
    *,
    model: str,
    temperature: float,
    reasoning_effort: str,
    messages: list,
    tools: list,
    tool_name: str
):
    if is_reasoning_model(model):
        responses_tools = [convert_chat_tool_to_responses_tool(t) for t in tools]

        resp = client.responses.create(
            model=model,
            reasoning={"effort": reasoning_effort},
            input=messages,
            tools=responses_tools,
            tool_choice={
                "type": "function",
                "name": tool_name
            },
            max_output_tokens=25000
        )

        return extract_response_tool_judgments(resp, tool_name)

    else:
        resp = client.chat.completions.create(
            model=model,
            temperature=temperature,
            messages=messages,
            tools=tools,
            tool_choice={
                "type": "function",
                "function": {"name": tool_name}
            }
        )

        return extract_tool_judgments(resp)


def call_single_agent(batch, temperature, model, reasoning_effort):
    user_content = json.dumps(batch, ensure_ascii=False)

    print(f"Calling Single Agent on transcript with {len(batch)} lines...")

    judgments = call_model_with_tool(
        model=model,
        temperature=temperature,
        reasoning_effort=reasoning_effort,
        messages=[
            {"role": "system", "content": judger_system_prompt},
            {"role": "user", "content": user_content}
        ],
        tools=judgment_tools,
        tool_name="provide_judgment"
    )

    judgments = normalize_items(judgments, batch)

    print(f"Single Agent returned {len(judgments)} judgments.")
    return judgments


def call_judgers(batch, temperature, model, reasoning_effort):
    results = []
    user_content = json.dumps(batch, ensure_ascii=False)

    for i, personality in enumerate(judger_personalities, start=1):
        print(f"Calling Judger {i} on transcript with {len(batch)} lines...")

        judgments = call_model_with_tool(
            model=model,
            temperature=temperature,
            reasoning_effort=reasoning_effort,
            messages=[
                {"role": "system", "content": personality},
                {"role": "user", "content": user_content}
            ],
            tools=judgment_tools,
            tool_name="provide_judgment"
        )

        judgments = normalize_items(judgments, batch)

        print(f"Judger {i} returned {len(judgments)} judgments.")

        results.append({
            "judger": i,
            "judgments": judgments
        })

    return results


def call_critic(batch, judger_outputs, temperature, model, reasoning_effort):
    user_content = json.dumps({
        "transcript": batch,
        "judger_outputs": judger_outputs
    }, ensure_ascii=False)

    print(f"Calling Critic on transcript with {len(batch)} lines...")

    judgments = call_model_with_tool(
        model=model,
        temperature=temperature,
        reasoning_effort=reasoning_effort,
        messages=[
            {"role": "system", "content": critic_system_prompt},
            {"role": "user", "content": user_content}
        ],
        tools=critic_tools,
        tool_name="aggregate_decisions"
    )

    judgments = normalize_items(judgments, batch)

    print(f"Critic returned {len(judgments)} judgments.")
    return judgments


# deprecated
# def call_batch_agent(batch, temp):
#     user_content = json.dumps(batch, ensure_ascii=False)

#     resp = client.chat.completions.create(
#         model=MODEL,
#         temperature=temp,
#         messages=[
#             {"role": "system", "content": SYSTEM_PROMPT},
#             {"role": "user", "content": user_content}
#         ]
#     )

#     text = resp.choices[0].message.content or ""
#     items = parse_jsonl(text)

#     expected = len(batch)
#     if len(items) != expected:
#         repair_prompt = f"""
# Convert the following model output into valid JSONL.
# Return exactly one JSON object per input line.
# Do not skip any lines.
# Schema:
# {{"lineNo": number, "decision": "yes" or "no", "spatialWordsDetected": string[], "justification": string}}

# Original model output:
# {text}

# Input lines:
# {user_content}
# """.strip()

#         repair_resp = client.chat.completions.create(
#             model=MODEL,
#             temperature=0,
#             messages=[
#                 {"role": "system", "content": "You convert outputs into valid JSONL only."},
#                 {"role": "user", "content": repair_prompt}
#             ]
#         )
#         repair_text = repair_resp.choices[0].message.content or ""
#         items = parse_jsonl(repair_text)

#     return normalize_items(items, batch)


def build_xlsx(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Spatial Results"

    ws.append([
        "line number",
        "participant",
        "utterance",
        "decision",
        "justification",
        "spatial words identified"
    ])

    for r in rows:
        ws.append([
            r["lineNo"],
            r.get("speaker", ""),
            r.get("utterance", ""),
            r.get("decision", ""),
            r.get("justification", ""),
            ", ".join(r.get("spatialWordsDetected", []))
        ])

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


@app.post("/analyze-xlsx")
async def analyze_xlsx(
    files: List[UploadFile] = File(...),
    temperature: float = Form(0),
    model: str = Form("gpt-5.5"),
    reasoning_effort: str = Form("medium"),
    agent_mode: str = Form("multi")
):
    final_rows = []

    for f in files:
        file_bytes = await f.read()
        name = f.filename.lower()

        if name.endswith(".xlsx"):
            parsed_rows = parse_xlsx(file_bytes)
        else:
            text = file_bytes.decode("utf-8", errors="ignore")

            if name.endswith(".cha"):
                parsed_rows = parse_cha(text)
            elif name.endswith(".csv"):
                parsed_rows = parse_csv(text)
            else:
                parsed_rows = parse_txt(text)

        if agent_mode == "single":
            judged = call_single_agent(
                parsed_rows,
                temperature,
                model,
                reasoning_effort
            )
        else:
            judger_outputs = call_judgers(
                parsed_rows,
                temperature,
                model,
                reasoning_effort
            )

            judged = call_critic(
                parsed_rows,
                judger_outputs,
                temperature,
                model,
                reasoning_effort
            )

        line_lookup = {row["lineNo"]: row for row in parsed_rows}

        for item in judged:
            src = line_lookup[item["lineNo"]]
            final_rows.append({
                "file": f.filename,
                "lineNo": src["lineNo"],
                "speaker": src["speaker"],
                "utterance": src["utterance"],
                "decision": item["decision"],
                "justification": item["justification"],
                "spatialWordsDetected": item["spatialWordsDetected"]
            })

    xlsx_file = build_xlsx(final_rows)

    return StreamingResponse(
        xlsx_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=spatial_results.xlsx"}
    )